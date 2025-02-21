import ttkbootstrap as tkb
from openpyxl import *
from tkinter import filedialog
import codecs
import csv

class Functions:
    def __init__(self):
        pass

    def choose_file(self):
        try:
            path_file = filedialog.askopenfilename()

            return path_file
        
        except OSError:
            print("Select a file")
    
    def csv_to_xlsx_velocity(self, file):

        self.xlsx_file = 'Punditlink Data.xlsx'

        wb = Workbook()
        ws = wb.active

        with codecs.open(file, 'r', encoding='utf-16-le') as csv_file:
            csv_reader = csv.reader(csv_file, delimiter='\t') 
            for row_idx, row in enumerate(csv_reader, start=1):
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(self.xlsx_file)

        return self.xlsx_file

    def organize_velocity_data(self):
        spreadsheet_path = self.choose_file()

        if spreadsheet_path.endswith('.csv'):

            file_xlsx = self.csv_to_xlsx_velocity(spreadsheet_path)

        
        wb = load_workbook(file_xlsx)

        if 'Sheet' in wb.sheetnames:
            ws = wb['Sheet']

            # Certifique-se de que "Organized Data" não exista previamente
            if "Organized Data" not in wb.sheetnames:
                wb.create_sheet("Organized Data", 1)

            ws1 = wb['Organized Data']

            lista_dados = []

            for row in ws.iter_rows(values_only=True):
                lista_dados_temporaria = list(row)
                lista_dados.append(lista_dados_temporaria)

            lista_valores = list(range(16, len(lista_dados), 25))

            ws1.append(lista_dados[14])

            for i, row_lista in enumerate(lista_dados):
                if i in lista_valores:
                    ws1.append(row_lista)

            wb.save(file_xlsx)

            
"""
if __name__ == "__main__":
    processor = Functions()
    processor.organize_velocity_data()
"""


